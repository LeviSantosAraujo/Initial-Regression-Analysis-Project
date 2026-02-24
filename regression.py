import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the raw data
df = pd.read_excel('Raw Data.xlsx')

# Filter for 2024 data
df_2024 = df[df['Year'] == 2024]

# Select target and features
target = 'Revenue ($)'
features = ['Units Sold', 'Market Share (%)', 'Regional 5G Coverage (%)', 
            '5G Subscribers (millions)', 'Avg 5G Speed (Mbps)', 'Preference for 5G (%)']

X = df_2024[features]
y = df_2024[target]

# Data summary
data_summary = df_2024[features + [target]].describe()

# Correlation with target
correlations = df_2024[features + [target]].corr()[target].drop(target)

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Create and train the linear regression model
model = LinearRegression()
model.fit(X_train, y_train)

# Make predictions on the test set
y_pred = model.predict(X_test)

# Evaluate the model
mse = mean_squared_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

# Create Excel report
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"
ws_data = wb.create_sheet("Data Summary")
ws_corr = wb.create_sheet("Correlations")
ws_model = wb.create_sheet("Model Results")

# Summary sheet
ws_summary['A1'] = "Regression Analysis Report - 2024 Data"
ws_summary['A1'].font = Font(size=16, bold=True)
ws_summary['A1'].alignment = Alignment(horizontal='center')
ws_summary.merge_cells('A1:E1')

ws_summary['A3'] = "Executive Summary"
ws_summary['A3'].font = Font(bold=True)
ws_summary['A4'] = "This report presents a linear regression analysis predicting Revenue ($) based on key 5G-related metrics for 2024 data."

ws_summary['A6'] = "Dataset Overview"
ws_summary['A6'].font = Font(bold=True)
ws_summary['A7'] = f"Source: Raw Data.xlsx"
ws_summary['A8'] = f"Filtered for Year: 2024"
ws_summary['A9'] = f"Total Samples: {len(df_2024)}"
ws_summary['A10'] = f"Target Variable: {target}"
ws_summary['A11'] = f"Features: {', '.join(features)}"
ws_summary['A12'] = f"Test Set Size: {len(X_test)} samples ({len(X_test)/len(df_2024)*100:.1f}% of data)"

ws_summary['A14'] = "Model Performance"
ws_summary['A14'].font = Font(bold=True)
ws_summary['A15'] = f"Mean Squared Error: {mse:.2f}"
ws_summary['A16'] = f"R-squared: {r2:.4f} (Explains {r2*100:.1f}% of variance in revenue)"

ws_summary['A18'] = "Intercept"
ws_summary['A18'].font = Font(bold=True)
ws_summary['B18'] = model.intercept_

ws_summary['A20'] = "Recommendations"
ws_summary['A20'].font = Font(bold=True)
ws_summary['A21'] = "Focus on high-impact features like Market Share and 5G Subscribers to boost revenue."
ws_summary['A22'] = "Monitor Regional 5G Coverage as it may negatively impact revenue."
ws_summary['A23'] = "Validate model with additional data or cross-validation for robustness."

# Data Summary sheet
ws_data['A1'] = "Data Summary"
ws_data['A1'].font = Font(size=14, bold=True)
for r, row in enumerate(dataframe_to_rows(data_summary.reset_index(), index=False, header=True), 1):
    for c, value in enumerate(row, 1):
        ws_data.cell(row=r, column=c, value=value)
        if r == 1:
            ws_data.cell(row=r, column=c).font = Font(bold=True)
            ws_data.cell(row=r, column=c).fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")

# Correlations sheet
ws_corr['A1'] = "Feature Correlations with Revenue"
ws_corr['A1'].font = Font(size=14, bold=True)
corr_df = correlations.to_frame().reset_index()
corr_df.columns = ['Feature', 'Correlation']
for r, row in enumerate(dataframe_to_rows(corr_df, index=False, header=True), 1):
    for c, value in enumerate(row, 1):
        ws_corr.cell(row=r, column=c, value=value)
        if r == 1:
            ws_corr.cell(row=r, column=c).font = Font(bold=True)
            ws_corr.cell(row=r, column=c).fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")

# Model Results sheet
ws_model['A1'] = "Model Coefficients"
ws_model['A1'].font = Font(size=14, bold=True)
ws_model['A2'] = "Feature"
ws_model['B2'] = "Coefficient"
ws_model['A2'].font = Font(bold=True)
ws_model['B2'].font = Font(bold=True)
ws_model['A2'].fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
ws_model['B2'].fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")

for i, (feature, coef) in enumerate(zip(features, model.coef_), 3):
    ws_model[f'A{i}'] = feature
    ws_model[f'B{i}'] = coef
    if coef > 0:
        ws_model[f'B{i}'].font = Font(color="00FF00")  # Green for positive
    else:
        ws_model[f'B{i}'].font = Font(color="FF0000")  # Red for negative

# Save the workbook
wb.save('regression_report.xlsx')

print("Detailed report saved to 'regression_report.xlsx'")
print("Chart saved to 'regression_plot.png'")
print("To send via email, run: python send_email.py")

# Plot actual vs predicted
plt.figure(figsize=(8, 6))
plt.scatter(y_test, y_pred, alpha=0.7)
plt.xlabel('Actual Revenue ($)')
plt.ylabel('Predicted Revenue ($)')
plt.title('Actual vs Predicted Revenue for 2024 Data')
plt.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], 'k--', lw=2)
plt.savefig('regression_plot.png')
plt.close()  # Close to avoid display issues